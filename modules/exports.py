"""
Export helpers for Tab 6: builds downloadable artifacts directly from a
fitted model's results — a betas-only time series, and a full Excel
workbook structured as Raw Variables -> Transformed Variables -> Betas
-> Contributions (Beta x Transformed), plus supporting sheets for
hyperparameters / ROI / synergy.

Design note: for every variable, Contribution = Beta * Transformed holds
exactly, by construction:
  - Intercept:           Raw = Transformed = 1.0
  - Media / Comp Media:  Transformed = Raw spend/impressions (what the
                          observation equation actually multiplies the
                          beta by — see kalman.py _build_observation_matrix;
                          carryover lives in beta's own decay, not in an
                          adstocked observation-side regressor)
  - Non-media / Comp Non-media / Price: Transformed = Raw (no transform
                          is applied to these — used as-is)
This keeps the identity Beta x Transformed = Contribution true for every
single column in the workbook, which is what makes the four-block layout
useful as a single coherent sheet rather than four unrelated tables.
"""

import io

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from modules.transforms import apply_transformation


def _build_variable_index(g):
    """Ordered (name, group, state_vector_index) for every modeled variable."""
    items = [("Intercept", "intercept", 0)]
    base = 1
    for i, col in enumerate(g["MEDIA_COLS"]):
        items.append((col, "media", base + i))
    base += g["N_MEDIA"]
    for j, col in enumerate(g["COMP_MEDIA_COLS"]):
        items.append((col, "comp_media", base + j))
    base += g["N_COMP"]
    for k, col in enumerate(g["OWN_NONMEDIA_COLS"]):
        items.append((col, "own_nonmedia", base + k))
    base += g["N_OWN_NONMEDIA"]
    for k, col in enumerate(g["COMP_NONMEDIA_COLS"]):
        items.append((col, "comp_nonmedia", base + k))
    base += g["N_COMP_NONMEDIA"]
    for p, col in enumerate(g["PRICE_COLS"]):
        items.append((col, "price", base + p))
    return items


def build_intercept_decomposition_df(res, df_full):
    """Period-by-period decomposition of the intercept's own state equation:

        I_t = G0 * I_(t-1)  +  Sum_k gamma_k * media_k,t^(n_k_intercept)

    Column order (left to right): Period, then for each intercept-effector
    channel a Media_<name> / Transformed_<name> / GammaXTransformed_<name>
    trio, then Intercept_Carryover (G0 * I_(t-1)) and Intercept_at_t (the
    full smoothed intercept level, I_t) as the last two columns.
    """
    g = res["g"]; params = res["params"]; x_smooth = res["x_smooth"]
    MEDIA_COLS = g["MEDIA_COLS"]
    TRANSFORM_TYPE = g["TRANSFORM_TYPE"]
    T = len(df_full)

    G0 = float(params["G0"])
    prev_intercept = np.empty(T)
    prev_intercept[1:] = x_smooth[:-1, 0]
    prev_intercept[0]  = x_smooth[0, 0]
    intercept_carryover = G0 * prev_intercept
    intercept_at_t = x_smooth[:, 0]

    data = {"Period": np.arange(T)}
    media_cols, trans_cols, gxt_cols = [], [], []

    for k, col in enumerate(g["INTERCEPT_EFFECTORS"]):
        ni_int = params["n_intercept"][k]
        raw = df_full[col].values.astype(float)
        if col in MEDIA_COLS:
            transformed = apply_transformation(
                raw, TRANSFORM_TYPE, ni_int, params["S_params"][MEDIA_COLS.index(col)])
        else:
            transformed = raw.copy()
        gxt = params["gamma"][k] * transformed

        media_name, trans_name, gxt_name = f"Media_{col}", f"Transformed_{col}", f"GammaXTransformed_{col}"
        data[media_name] = raw
        data[trans_name] = transformed
        data[gxt_name]   = gxt
        media_cols.append(media_name); trans_cols.append(trans_name); gxt_cols.append(gxt_name)

    data["Intercept_Carryover"] = intercept_carryover
    data["Intercept_at_t"] = intercept_at_t

    ordered = ["Period"] + media_cols + trans_cols + gxt_cols + ["Intercept_Carryover", "Intercept_at_t"]
    return pd.DataFrame(data)[ordered]


def build_betas_df(res, df_full):
    """Period + one Beta_<variable> column per modeled variable."""
    g = res["g"]; x_smooth = res["x_smooth"]
    items = _build_variable_index(g)
    out = pd.DataFrame({"Period": np.arange(len(df_full))})
    for name, _group, idx in items:
        out[f"Beta_{name}"] = x_smooth[:, idx]
    return out


def _raw_and_transformed(name, group, df_full, adstocked_media, T):
    if group == "intercept":
        return np.ones(T), np.ones(T)
    raw_vals = df_full[name].values.astype(float)
    if group in ("media", "comp_media"):
        # Transformed = Raw here too: the observation equation multiplies
        # beta by raw spend/impressions, not adstocked media.
        return raw_vals, raw_vals.copy()
    return raw_vals, raw_vals.copy()


def build_model_data_df(res, config, df_full):
    """Single-sheet layout: Period, Target, then Raw_* | Transformed_* |
    Beta_* | Contribution_* blocks, in that left-to-right order."""
    g = res["g"]; x_smooth = res["x_smooth"]; adstocked_media = res["adstocked_media"]
    items = _build_variable_index(g)
    T = len(df_full)

    data = {"Period": np.arange(T), f"Target_{config['target']}": df_full[config["target"]].values}
    raw_cols, trans_cols, beta_cols, contrib_cols = [], [], [], []

    for name, group, idx in items:
        raw_vals, trans_vals = _raw_and_transformed(name, group, df_full, adstocked_media, T)
        beta_vals = x_smooth[:, idx]
        contrib_vals = beta_vals * trans_vals

        raw_name, trans_name = f"Raw_{name}", f"Transformed_{name}"
        beta_name, contrib_name = f"Beta_{name}", f"Contribution_{name}"
        data[raw_name] = raw_vals
        data[trans_name] = trans_vals
        data[beta_name] = beta_vals
        data[contrib_name] = contrib_vals
        raw_cols.append(raw_name); trans_cols.append(trans_name)
        beta_cols.append(beta_name); contrib_cols.append(contrib_name)

    ordered = ["Period", f"Target_{config['target']}"] + raw_cols + trans_cols + beta_cols + contrib_cols
    return pd.DataFrame(data)[ordered]


_BLOCK_COLORS = {
    "Raw_":          "3B82F6",  # blue   — matches the app's info-box accent
    "Transformed_":  "16A34A",  # green  — matches the prophet/positive-beta accent
    "Beta_":         "F97316",  # orange — matches the per-channel-bounds accent
    "Contribution_": "9333EA",  # purple — matches the nevergrad accent
}
_NEUTRAL_HEADER = "334155"

_INTERCEPT_BLOCK_COLORS = {
    "Media_":                "3B82F6",  # blue   — same convention as Raw_ above
    "Transformed_":          "16A34A",  # green  — same convention as Model_Data
    "GammaXTransformed_":    "9333EA",  # purple — same convention as Contribution_
    "Intercept_Carryover":   "F97316",  # orange — the G0 * I_(t-1) piece
    "Intercept_at_t":        "DC2626",  # red    — the resulting full intercept level
}


def _style_intercept_decomp_sheet(ws, columns):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    data_font   = Font(name="Arial")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        fill_color = _NEUTRAL_HEADER
        for prefix, color in _INTERCEPT_BLOCK_COLORS.items():
            if col_name == prefix or col_name.startswith(prefix):
                fill_color = color
                break
        cell.fill = PatternFill("solid", fgColor=fill_color)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
    ws.freeze_panes = "B2"


def _style_model_data_sheet(ws, columns):
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    data_font   = Font(name="Arial")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        fill_color = _NEUTRAL_HEADER
        for prefix, color in _BLOCK_COLORS.items():
            if col_name.startswith(prefix):
                fill_color = color
                break
        cell.fill = PatternFill("solid", fgColor=fill_color)
        ws.column_dimensions[get_column_letter(col_idx)].width = max(14, len(col_name) + 2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
    ws.freeze_panes = "C2"


def _style_simple_sheet(ws):
    header_font = Font(name="Arial", bold=True)
    data_font   = Font(name="Arial")
    for cell in ws[1]:
        cell.font = header_font
        ws.column_dimensions[cell.column_letter].width = max(14, len(str(cell.value or "")) + 2)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
    ws.freeze_panes = "A2"


def build_master_workbook_bytes(res, config, df_full):
    """Full .xlsx: Model_Data (Raw -> Transformed -> Beta -> Contribution),
    Hyperparameters, ROI, Synergy (if configured), and a Legend sheet."""
    model_data_df = build_model_data_df(res, config, df_full)
    intercept_decomp_df = build_intercept_decomposition_df(res, df_full)
    param_df = res["param_df"]
    roi_df   = res["roi_df"]
    synergy_df = res.get("synergy_df")
    has_synergy = synergy_df is not None and not synergy_df.empty

    legend_df = pd.DataFrame({
        "Block": ["Raw_*", "Transformed_*", "Beta_*", "Contribution_*",
                  "Media_* (Intercept_Decomposition)", "Transformed_* (Intercept_Decomposition)",
                  "GammaXTransformed_*", "Intercept_Carryover", "Intercept_at_t"],
        "Meaning": [
            "The variable's original value, as uploaded / merged into the dataset.",
            "What the model actually multiplies the beta by. For media and "
            "competitor media this is the adstocked (carry-over adjusted) spend. "
            "For every other variable type, no transform is applied, so this "
            "equals Raw.",
            "The EKF's smoothed, time-varying coefficient for that variable.",
            "Beta x Transformed for that variable and period — its modeled "
            "contribution to the target KPI in that period.",
            "Raw value of an intercept-effector channel (e.g. SEO spends, TV Grp) "
            "for that period — same underlying number as Model_Data's Raw_*, "
            "repeated here for readability.",
            "The intercept-effector's raw value run through its own n_intercept "
            "transform (Hill/power, per Transform Type), i.e. media_k,t^n_k_intercept.",
            "gamma_k * Transformed_* — that channel's boost contribution to the "
            "intercept's state equation for that period.",
            "G0 * Intercept_(t-1) — the persisted/carried-over piece of the "
            "intercept's state equation.",
            "The full smoothed intercept level, I_t = Intercept_Carryover + "
            "Sum_k GammaXTransformed_k (up to Kalman process noise / baseline "
            "flooring) — this is what enters the observation equation as-is.",
        ],
    })

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        model_data_df.to_excel(writer, sheet_name="Model_Data", index=False)
        intercept_decomp_df.to_excel(writer, sheet_name="Intercept_Decomposition", index=False)
        param_df.to_excel(writer, sheet_name="Hyperparameters", index=False)
        roi_df.to_excel(writer, sheet_name="ROI", index=False)
        if has_synergy:
            synergy_df.to_excel(writer, sheet_name="Synergy", index=False)
        legend_df.to_excel(writer, sheet_name="Legend", index=False)

        wb = writer.book
        _style_model_data_sheet(wb["Model_Data"], model_data_df.columns)
        _style_intercept_decomp_sheet(wb["Intercept_Decomposition"], intercept_decomp_df.columns)
        _style_simple_sheet(wb["Hyperparameters"])
        _style_simple_sheet(wb["ROI"])
        if has_synergy:
            _style_simple_sheet(wb["Synergy"])
        legend_ws = wb["Legend"]
        _style_simple_sheet(legend_ws)
        for row_idx in range(2, legend_ws.max_row + 1):
            legend_ws.column_dimensions["B"].width = 90
            legend_ws.cell(row=row_idx, column=2).alignment = Alignment(wrap_text=True)

    buffer.seek(0)
    return buffer.getvalue()
